from openpyxl import Workbook

from typing import List, Union, Iterator, Tuple
from bs4.element import Tag


class Parser:
    def __init__(self, file_path):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.file_path = file_path
        self.load_workbook()

    def load_workbook(self):

        '''
            reads file and loads it into respective format
            ready for dumping to file
        '''
        raise NotImplemented

    def _read_file(self):
        """
        returns the data contained in a file
        Returns
        -------
        data : bytes
                File stream
        """
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                data = f.read()
            return data
        except:
            raise Exception("Error while reading input file")

    def _get_row(self, table: Tag, tags: Union[List, str]) -> Iterator[Tag]:
        '''
            reads all tags present inside a parent tag
            and returns a generator
            Parameters
            ----------
            table : Tag
                    parent tag
            tags : List[str] or str
                    tags to search for in parent tag `table`
            Returns
            -------
            iter_tag : Iterator[Tag]
                    Generator consisting of all occurences of `tags` in `table`
        '''
        row_data = table.find_all(tags)
        for each in row_data:
            yield each

    def _pre_validate_and_format(self, start_row: int, start_col: int, col: Tag) -> str:
        """
            formats cells according to attribute tags/ metadata
            Parameters
            ----------
            start_row : int
                    Start of the row
            start_col : int
                    Start of the column
            col : Tag
                    Cell details including value and metadata
            Returns
            -------
            value: str
                    Cell value
        """
        attrs = col.attrs
        end_row = start_row
        end_col = start_col
        if "colspan" in attrs:
            colspan = int(attrs.get("colspan", 1))
            end_col += colspan - 1
        if "rowspan" in attrs:
            rowspan = int(attrs.get("rowspan", 1))
            end_row += rowspan - 1

        self.ws.merge_cells(start_row=start_row, end_row=end_row,
                            start_column=start_col, end_column=end_col)
        

        # TODO: Handle bold, italics and other attributes
        return col.text.strip()
            

    def _write_cell(self, row: int, col: int, val: str) -> None:
        """
            writes value to cell
            Parameters
            ----------
            row : int
                    row number
            col : int
                    column number
            val : str
                    Value to write in cell
        """
        self.ws.cell(row=row, column=col).value = val

    def get_workbook(self) -> Workbook:
        return self.wb

    def _save_workbook(self, loc: str) -> bool:
        """
            saves workbook to specified location
            Parameters
            ----------
            loc : str
                    save location for workbook
        """
        try:
            self.wb.save(loc)
            return True
        except:
            return False
