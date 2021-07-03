from openpyxl import Workbook
from bs4 import BeautifulSoup
from html2excel.base.parser import BaseParser

from bs4.element import Tag
from collections import defaultdict

from typing import Dict, List, Tuple, Set, Optional


class ExcelParser(BaseParser):
    def __init__(self, file_path: str, enc: str = 'utf-8'):
        """
        Parameters
        ----------
        file_path : str
                Path where the html file is located
        
        enc: str, optional
                Encoding to use while reading file
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        super().__init__(file_path, enc)

    def get_workbook(self) -> Workbook:
        return self.wb

    def _save_workbook(self, loc: str) -> bool:
        """
            saves workbook to specified location
            Parameters
            ----------
            loc : str
                    save location for workbook
        """
        try:
            self.wb.save(loc)
            return True
        except:
            return False

    def _write_cell(self, row: int, col: int, val: str) -> None:
        """
            writes value to cell
            Parameters
            ----------
            row : int
                    row number
            col : int
                    column number
            val : str
                    Value to write in cell
        """
        self.ws.cell(row=row, column=col).value = val

    

    

    def _pre_validate_and_format(self, start_row: int, start_col: int, col: Tag) -> str:
        """
        formats cells according to attribute tags/ metadata
        Parameters
        ----------
        start_row : int
                Start of the row
        start_col : int
                Start of the column
        col : Tag
                Cell details including value and metadata
        Returns
        -------
        value: str
                Cell value
        """
        attrs = col.attrs
        end_row = start_row
        end_col = start_col
        if "colspan" in attrs:
            colspan = int(attrs.get("colspan", 1))
            end_col += colspan - 1
        if "rowspan" in attrs:
            rowspan = int(attrs.get("rowspan", 1))
            end_row += rowspan - 1

        # Merge cells
        self.ws.merge_cells(
            start_row=start_row,
            end_row=end_row,
            start_column=start_col,
            end_column=end_col,
        )

        # TODO: Handle bold, italics and other attributes
        return col.text.strip()

    def load_workbook(self):
        data = self._read_file()
        soup = BeautifulSoup(data, features="html5lib")

        all_data_html = soup.html.body.find_all(recursive=False)
        if all_data_html is None:
            raise Exception("No table found")

        cell_map_dict = self.get_cell_value_map(all_data_html)
        for row in cell_map_dict:
            for col, tag in cell_map_dict[row]:
                cell_value = self._pre_validate_and_format(row, col, tag)
                self._write_cell(row, col, cell_value)

    def to_excel(self, save_file_path: str) -> None:
        """
        convert html file to excel and save it to a path
        Parameters
        ----------
        save_file_path : str
                file path where the excel file is saved
        """
        self.load_workbook()
        self._save_workbook(save_file_path)
