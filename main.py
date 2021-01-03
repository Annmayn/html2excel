from openpyxl import Workbook
from bs4 import BeautifulSoup
import re

from typing import List, Union, Iterator, Tuple
from bs4.element import Tag

class Parser:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.input_file = "./html_file.html"
        self.table_data_regex = "(<table(?:.*\n)*?.*?</table>)"
        self.row_data_regex = "(<tr.*?>(?:.*\n)*?.*?<\/tr>)"
        self.cell_data_regex = "<t[hd].*?>((?:.*\n)*?.*?)<\/t[hd]>"
        self.cell_attribute_regex = "<t[hd]().*?)>(?:.*\n)*?.*?<\/t[hd]>"
        self.data = ''
        self.soup = BeautifulSoup()

    def read_file(self):
        try:
            with open(self.input_file, 'r') as f:
                self.data = f.read()
                self.soup = BeautifulSoup(self.data, features="html.parser")
        except:
            raise Exception("Error while reading input file")

    @classmethod
    def get_table_data(cls, soup: BeautifulSoup) -> Union[Tag, None]:
        # def get_table_data(cls, pattern: str, data: str, join: str = '-') -> str:
        # table_data = re.search(pattern, data)
        # table_data = join.join(table_data.groups()) if table_data is not None else ''
        # return table_data
        return soup.table
    
    def get_row(self, table: Tag, tags: Union[List, str]) -> Iterator[Tag]:
        # def get_row(cls, pattern: str, data: str) -> List:
        # row_data = re.findall(pattern, data)
        row_data = table.find_all(tags)
        for each in row_data:
            yield each

    @classmethod
    def strip_tags(cls, val: str) -> str:
        strip_regex = "<.*?>"
        res = re.sub(strip_regex, '', val)
        return res.strip()

    def pre_validate_and_format(self, i: int, j: int, col: Tag) -> Tuple[int, str]:
        attrs = col.attrs
        end = j
        if "colspan" in attrs:
            colspan = int(attrs.get("colspan", 1))
            end += colspan - 1
            self.ws.merge_cells(start_row=i, end_row=i, start_column=j, end_column=end)
        end += 1
        return (end, col.text.strip())

    
    def write_cell(self, row, col, val) -> None:
        self.ws.cell(row=row, column=col).value = val
    
    def get_workbook(self) -> Workbook:
        return self.wb

    def save_workbook(self, loc) -> bool:
        try:
            self.wb.save(loc)
            return True
        except:
            return False


parser = Parser()
parser.read_file()
# table_data = parser.get_table_data(parser.table_data_regex, parser.data)
# data_rows = parser.get_row(parser.row_data_regex, table_data)

# for i, row in enumerate(data_rows, 1):
#     # if i == 1:
#     #     print(row)
#     columns = parser.get_row(parser.cell_data_regex, row)
#     cell_attributes = parser.get_row(parser.cell_data_regex, row)
#     for j, col in enumerate(columns, 1):
#         # if j==1:
#         #     print('COL: ', col)
#         col = parser.strip_tags(col)
#         parser.write_cell(i, j, col)

# parser.save_workbook("./trial.xlsx")

table_data = parser.get_table_data(parser.soup)
data_rows = parser.get_row(table_data, ["tr"])
for i, row in enumerate(data_rows, 1):
    columns = parser.get_row(row, ["th", "td"])
    next_j = 1
    for j, col in enumerate(columns, 1):
        j = next_j
        # if i==j==1:
        #     print("Col->", col, col.attrs)
        next_j, col_data = parser.pre_validate_and_format(i, j, col)
        parser.write_cell(i, j, col_data)

parser.save_workbook("./trial.xlsx")