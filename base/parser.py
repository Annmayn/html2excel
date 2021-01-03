from openpyxl import Workbook
from bs4 import BeautifulSoup
import re

from typing import List, Union, Iterator, Tuple
from bs4.element import Tag


class Parser:
    def __init__(self, file_path):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.file_path = file_path
        self.table_data_regex = "(<table(?:.*\n)*?.*?</table>)"
        self.row_data_regex = "(<tr.*?>(?:.*\n)*?.*?<\/tr>)"
        self.cell_data_regex = "<t[hd].*?>((?:.*\n)*?.*?)<\/t[hd]>"
        self.cell_attribute_regex = "<t[hd]().*?)>(?:.*\n)*?.*?<\/t[hd]>"
        self.data = ''

    def read_file(self):
        try:
            with open(self.file_path, 'r') as f:
                data = f.read()
            return data
        except:
            raise Exception("Error while reading input file")

    @classmethod
    def get_table_data(cls, soup: BeautifulSoup) -> Union[Tag, None]:
        return soup.table

    def get_row(self, table: Tag, tags: Union[List, str]) -> Iterator[Tag]:
        # def get_row(cls, pattern: str, data: str) -> List:
        # row_data = re.findall(pattern, data)
        row_data = table.find_all(tags)
        for each in row_data:
            yield each

    @classmethod
    def strip_tags(cls, val: str) -> str:
        strip_regex = "<.*?>"
        res = re.sub(strip_regex, '', val)
        return res.strip()

    def pre_validate_and_format(self, i: int, j: int, col: Tag) -> Tuple[int, str]:
        attrs = col.attrs
        end = j
        if "colspan" in attrs:
            colspan = int(attrs.get("colspan", 1))
            end += colspan - 1
            self.ws.merge_cells(start_row=i, end_row=i,
                                start_column=j, end_column=end)
        # TODO: Handle bold, italics and other attributes
        end += 1
        return (end, col.text.strip())

    def write_cell(self, row, col, val) -> None:
        self.ws.cell(row=row, column=col).value = val

    def get_workbook(self) -> Workbook:
        return self.wb

    def save_workbook(self, loc) -> bool:
        try:
            self.wb.save(loc)
            return True
        except:
            return False
